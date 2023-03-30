import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
from django.shortcuts import render
from django.http import FileResponse


name_excel = 'name.xlsx' #Название выходного файла
##__________Сортировка_____________##
def sort(sheet):
    for run in range(sheet.max_row):
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=6).value is None:
                for column in range(1, 7):
                    sheet.cell(row=i, column=column).value, sheet.cell(row=i + 1,column=column).value = sheet.cell(
                        row=i + 1, column=column).value, sheet.cell(row=i, column=column).value
            elif sheet.cell(row=i + 1, column=6).value is None:
                pass
            else:
                if sheet.cell(row=i, column=6).value < sheet.cell(row=i + 1, column=6).value:
                    for column in range(1, 7):
                        sheet.cell(row=i, column=column).value, sheet.cell(row=i + 1,column=column).value = sheet.cell(
                            row=i + 1, column=column).value, sheet.cell(row=i, column=column).value
##__________Создаем шапку__________##
def shapka(sheet):
    sheet.insert_rows(idx = 1, amount = 2)
    ##                  Заполняем ячейки                   ##
    sheet.cell(row=1, column=1).value = 'Филиал'
    sheet.cell(row=1, column=2).value = 'Сотрудник'
    sheet.cell(row=1, column=3).value = 'Налоговая база'
    sheet.cell(row=1, column=4).value = 'Налог'
    sheet.cell(row=2, column=4).value = 'Исчислено всего'
    sheet.cell(row=2, column=5).value = 'Исчислено всего по формуле'
    sheet.cell(row=1, column=6).value = 'Отклонения'
    ##                  Объединения                   ##
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('F1:G2')
    sheet.merge_cells('D1:E1')
    ##                  Стили                   ##
    double = Side(border_style="medium", color="9EB6B1")
    for j in range(2):
        for i in range(7):
            sheet.cell(row=j + 1, column=i + 1).font = Font(bold=True, color='010775', size=10, name='Arial')
            sheet.cell(row=j + 1, column=i + 1).fill = PatternFill('solid', fgColor="c9e6e4")
            sheet.cell(row=j + 1, column=i + 1).alignment = Alignment(horizontal='center', vertical="center",
                                                                       wrapText=True, shrinkToFit=True)
            sheet.cell(row=j + 1, column=i + 1).border = Border(top=double, bottom=double, left=double, right=double)
    ##                  Ширина/высота колонок                   ##
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 37
    sheet.column_dimensions['C'].width = 13
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['E'].width = 17
    sheet.row_dimensions[2].height = 25


def index(request):
    if request.POST:
        ##      Загружаем файл и создаем свой       ##
        file = request.FILES['input_file']
        book = openpyxl.Workbook()
        sheet = book.active
        book_input = openpyxl.load_workbook(file)
        sheet_input = book_input.active
        ##                                Парсим sheet_input и собираем sheet                              ##
        for row in range(1, sheet_input.max_row - 2):
            row_sheet_input = row + 2
            sheet.cell(row=row, column=1).value = sheet_input.cell(row=row_sheet_input, column=1).value
            sheet.cell(row=row, column=2).value = sheet_input.cell(row=row_sheet_input, column=2).value
            sheet.cell(row=row, column=3).value = sheet_input.cell(row=row_sheet_input, column=5).value
            sheet.cell(row=row, column=4).value = sheet_input.cell(row=row_sheet_input, column=6).value
            try:
                if float(sheet_input.cell(row=row_sheet_input, column=5).value) < 5000000:
                    sheet.cell(row=row, column=5).value = sheet_input.cell(row=row_sheet_input, column=5).value * 0.13
                else:
                    sheet.cell(row=row, column=5).value = sheet_input.cell(row=row_sheet_input, column=5).value * 0.15
                sheet.cell(row=row, column=6).value = sheet_input.cell(row=row_sheet_input, column=6).value - sheet.cell(row=row,
                                                                                                               column=5).value
            except:
                sheet.cell(row=row, column=5).value = None
                sheet.cell(row=row, column=6).value = None
        ##                                Сортируем и добавляем шапку                              ##
        sort(sheet=sheet)
        shapka(sheet=sheet)
        ##                                Красим ячейки по условию                              ##
        for row in range(3, sheet.max_row + 1):
            if sheet.cell(row=row, column=6).value == 0:
                sheet.cell(row=row, column=6).fill = PatternFill('solid', fgColor="60f542")
            else:
                sheet.cell(row=row, column=6).fill = PatternFill('solid', fgColor="f70202")
            sheet.merge_cells(f'F{row}:G{row}')
        ##                                Отправляем файл пользователю                              ##
        book.save(name_excel)
        return FileResponse(open(name_excel,'rb'))
    return render(request, 'html_form/index.html')
